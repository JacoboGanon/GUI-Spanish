# Imports
from tkinter import Tk, Label, Frame, Entry, Button, font, Listbox, messagebox, END, IntVar
from tkinter.ttk import Combobox, Checkbutton
import openpyxl
from datetime import datetime
import json

# Create Window
root = Tk()
root.geometry('%dx%d' % (root.winfo_screenwidth(), root.winfo_screenheight()))

# Main Loop
class FrontEnd:
    def __init__(self):
        # Set default font
        self.defaultFont = font.nametofont("TkDefaultFont")
        # Create 3 payment types
        self.payment_types = ('Tarjeta', 'Efectivo', 'Transferencia')
        # Open Workbook
        self.wb = openpyxl.load_workbook('income_expenses.xlsx')

        # Overriding default-font with custom settings
        # i.e changing font-family, size and weight
        self.defaultFont.configure(family="Times", size=20)
        # Heights (Format = self.height, number of columns, number in column)
        self.total_price = 0
        self.selected_payment_type = 0
        self.height1_1 = 0.5
        self.height2_1 = .33
        self.height2_2 = .67
        self.height3_1 = .25
        self.height3_2 = .5
        self.height3_3 = .75
        self.height4_1 = .2
        self.height4_2 = .4
        self.height4_3 = .6
        self.height4_4 = .8
        # Widths (Format = self.width, number of rows, number in row)
        self.width1_1 = .5
        self.width2_1 = 0.33
        self.width2_2 = 0.67
        self.width3_1 = .25
        self.width3_2 = .5
        self.width3_3 = .75
        self.width4_1 = .2
        self.width4_2 = .4
        self.width4_3 = .6
        self.width4_4 = .8

        # Colors
        self.WHITE = '#242B2E'
        self.ORANGE = "#CAD5E2"
        self.GREY = "#6C36AD"
        try:
            self.complete_products_data = json.load(open('products_information.txt', 'r'))
        except:
            self.complete_products_data = []

        try:
            self.complete_students_data = json.load(open('Students.txt', 'r'))
        except:
            self.complete_students_data = []
        try:
            self.complete_classes_data = json.load(open('Available_Classes.txt', 'r'))
        except:
            self.complete_classes_data = []

        # Labels presented
        self.list_of_labels = []

        # Listbox Selections
        self.listbox_selection = 0

    def remove_everything(self):
        # Remove 'Clases'
        try:
            self.remove_student_button.destroy()
            self.add_student_button.destroy()
            self.class_frame1.destroy()
            self.class_frame2.destroy()
        except AttributeError:
            pass
        # Remove 'Añadir clase, clase'
        try:
            self.list_of_activities.destroy()
            self.day_of_week.destroy()
            self.time_schedule_entry.destroy()
            self.teacher_entry.destroy()
            self.available_seats_entry.destroy()
            self.time_schedule_label.destroy()
            self.teacher_label.destroy()
            self.available_seats_label.destroy()
            self.frame1_class_page.destroy()
            self.frame2_class_page.destroy()
            self.frame3_class_page.destroy()
            self.frame4_class_page.destroy()
            self.frame5_class_page.destroy()
            self.frame6_class_page.destroy()
            self.frame7_class_page.destroy()
        except AttributeError:
            pass
        # Remove 'Añadir Actividad'
        try:
            self.add_activity_button.destroy()
            self.add_activity_entry.destroy()
            self.add_levels_label.destroy()
            self.add_levels_entry.destroy()
            self.frame1_activity_page.destroy()
            self.frame2_activity_page.destroy()
            self.frame3_activity_page.destroy()
        except AttributeError:
            pass
        # Remove 'Añadir clase'
        try:
            self.activity_button.destroy()
            self.add_class_button.destroy()
            self.frame3_admin_page.destroy()
            self.frame4_admin_page.destroy()
        except AttributeError:
            pass
        # Remove Sell
        try:
            self.class_button.destroy()
            self.product_button.destroy()
            self.frame1_seller.destroy()
            self.frame2_seller.destroy()
        except AttributeError:
            pass
        # Remove self
        try:
            self.admin_mode.destroy()
            self.seller_mode.destroy()
            self.frame1_sell.destroy()
            self.frame2_sell.destroy()
        except AttributeError:
            pass
        # Remove Admin password
        try:
            self.password_entry.destroy()
            self.password_button.destroy()
            self.frame1_admin_password.destroy()
            self.frame2_admin_password.destroy()
        except AttributeError:
            pass
        # Remove 'Productos'
        try:
            self.item_id_label.destroy()
            self.item_id_entry.destroy()
            self.item_to_sell_entry.destroy()
            self.item_to_sell_label.destroy()
            self.item_provider_label.destroy()
            self.item_provider_entry.destroy()
            self.price_entry.destroy()
            self.price_label.destroy()
            self.current_inventory_entry.destroy()
            self.current_inventory_label.destroy()
            self.cost_entry.destroy()
            self.cost_label.destroy()
            self.register_button.destroy()
            self.frame3_admin_page.destroy()
            self.frame4_admin_page.destroy()
            self.frame5_admin_page.destroy()
            self.frame6_admin_page.destroy()
            self.frame7_admin_page.destroy()
            self.frame8_admin_page.destroy()
            self.frame9_admin_page.destroy()
            self.frame10_admin_page.destroy()
        except AttributeError:
            pass
        # Remove 'Añadir Inverntario'
        try:

            self.listbox_of_products_id.destroy()
            self.add_inventory_entry.destroy()
            self.add_inventory_button.destroy()
            self.frame3_admin_page.destroy()
            self.frame4_admin_page.destroy()
            self.frame5_admin_page.destroy()
            self.frame6_admin_page.destroy()
            self.frame7_admin_page.destroy()
        except AttributeError:
            pass
        # Remove 'Quitar Producto'
        try:
            self.listbox_of_products_id.destroy()
            self.remove_product_button.destroy()
            self.frame3_admin_page.destroy()
            self.frame4_admin_page.destroy()
        except AttributeError:
            pass

    def main_page(self):
        self.remove_everything()
        self.return_button.destroy()
        # Remove admin page
        try:
            self.choose_to_add.destroy()
            self.listbox_button.destroy()
            self.frame1_admin_page.destroy()
            self.frame2_admin_page.destroy()
        except AttributeError:
            pass
        self.frame1_sell = Frame(root, bg=self.GREY)
        self.frame1_sell.place(relx=self.width2_1 - .15, rely=self.height1_1 - .04, relwidth=.3, relheight=.08)
        self.frame2_sell = Frame(root, bg=self.GREY)
        self.frame2_sell.place(relx=self.width2_2 - .15, rely=self.height1_1 - .04, relwidth=.3, relheight=.08)
        self.admin_mode = Button(self.frame1_sell, text="Administrar", command=self.admin_ask_password, bg=self.ORANGE, fg=self.WHITE)
        self.admin_mode.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        self.seller_mode = Button(self.frame2_sell, text="Vender", command=self.sell_mode, bg=self.ORANGE, fg=self.WHITE)
        self.seller_mode.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def admin_ask_password(self):
        self.remove_everything()
        self.return_button = Button(root, text="Regresar", command=self.main_page, bg=self.ORANGE, fg=self.WHITE)
        self.return_button.place(relwidth=0.13, relheight=0.06)
        # Create frames
        self.frame1_admin_password = Frame(root, bg=self.GREY)
        self.frame1_admin_password.place(relx=self.width2_1 - .15, rely=self.height1_1 - .03, relwidth=.3, relheight=.06)
        self.frame2_admin_password = Frame(root, bg=self.GREY)
        self.frame2_admin_password.place(relx=self.width2_2 - .15, rely=self.height1_1 - .03, relwidth=.3, relheight=.06)
        # Create password entry and button
        self.password_entry = Entry(self.frame1_admin_password, show='*', font=20)
        self.password_entry.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        self.password_button = Button(self.frame2_admin_password, text="Registrar Contraseña", command=self.get_password, bg=self.ORANGE, fg=self.WHITE)
        self.password_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def get_password(self):
        # Confirm password is correct
        attempted_password = self.password_entry.get()
        password = "Flicflac2021!"
        if attempted_password == password:
            self.admin_page()

    def admin_page(self):
        self.remove_everything()
        # Create Frames
        self.frame1_admin_page = Frame(root, bg=self.GREY)
        self.frame1_admin_page.place(relx=self.width2_1 - .15, rely=self.height1_1 - .1, relwidth=.3, relheight=.2)
        self.frame2_admin_page = Frame(root, bg=self.GREY)
        self.frame2_admin_page.place(relx=self.width2_2 - .15, rely=self.height1_1 - .05, relwidth=.3, relheight=.1)
        # Create Listbox
        self.choose_to_add = Listbox(self.frame1_admin_page, height=4)
        self.choose_to_add.insert(1, 'Añadir Clase')
        self.choose_to_add.insert(2, 'Añadir Producto')
        self.choose_to_add.insert(3, 'Añadir Inventario')
        self.choose_to_add.insert(4, 'Quitar Producto')
        self.choose_to_add.insert(5, 'Añadir Egreso')
        self.choose_to_add.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        # Create button to get listbox information
        self.listbox_button = Button(self.frame2_admin_page, text="Seleccionar", command=self.gateway_to_decision, bg=self.ORANGE, fg=self.WHITE)
        self.listbox_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def gateway_to_decision(self):
        if self.choose_to_add.get(self.choose_to_add.curselection()) == "Añadir Clase":
            self.add_class_to_sell()
        elif self.choose_to_add.get(self.choose_to_add.curselection()) == "Añadir Producto":
            self.add_product_to_sell()
        elif self.choose_to_add.get(self.choose_to_add.curselection()) == "Añadir Inventario":
            self.inventory_list()
        elif self.choose_to_add.get(self.choose_to_add.curselection()) == 'Quitar Producto':
            self.remove_product_form()
        elif self.choose_to_add.get(self.choose_to_add.curselection()) == 'Añadir Egreso':
            self.add_expense()

    def add_expense(self):
        # Remove everything
        self.remove_everything()

        # Move 2 frames
        self.frame1_admin_page.place_forget()
        self.frame2_admin_page.place_forget()
        self.frame1_admin_page.place(rely=self.height3_1-.04, relx=self.width2_1-.15, relwidth=.3, relheight=.08)
        self.frame2_admin_page.place(rely=self.height3_1-.04, relx=self.width2_2-.15, relwidth=.3, relheight=.08)

        # Create 5 frames
        self.frame3_admin_page = Frame(root, bg=self.GREY)
        self.frame4_admin_page = Frame(root, bg=self.GREY)
        self.frame5_admin_page = Frame(root, bg=self.GREY)
        self.frame6_admin_page = Frame(root, bg=self.GREY)
        self.frame7_admin_page = Frame(root, bg=self.GREY)

        # Place 5 frames
        self.frame3_admin_page.place(relx=self.width3_1 - .1, rely=self.height3_2-.04, relwidth=.2, relheight=.08)
        self.frame4_admin_page.place(relx=self.width3_2 - .1, rely=self.height3_2 - .04, relwidth=.2, relheight=.08)
        self.frame5_admin_page.place(relx=self.width3_3 - .1, rely=self.height3_2 - .04, relwidth=.2, relheight=.08)
        self.frame6_admin_page.place(relx=self.width2_1 - .15, rely=self.height3_3 - .04, relwidth=.3, relheight=.08)
        self.frame7_admin_page.place(relx=self.width2_2 - .15, rely=self.height3_3 - .04, relwidth=.3, relheight=.08)

        # Payment type listbox
        self.payment_type = Listbox(self.frame3_admin_page, height=3)
        self.payment_type.bind('<<ListboxSelect>>', self.get_value2)
        self.payment_type.place(relx=.05, rely=.1, relheight=.8, relwidth=.9)
        for i in range(len(self.payment_types)):
            self.payment_type.insert(i, self.payment_types[i])

        # Create Labels
        self.amount_paid_label = Label(self.frame4_admin_page, text='Cantidad Pagada', bg=self.GREY)
        self.category_label = Label(self.frame5_admin_page, text='Categoria', bg=self.GREY)
        self.description_label = Label(self.frame6_admin_page, text='Cantidad Pagada', bg=self.GREY)

        # Place Labels
        self.amount_paid_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.category_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.description_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)

        # Create Entries
        self.amount_paid_entry = Entry(self.frame4_admin_page, font=20)
        self.category_entry = Entry(self.frame5_admin_page, font=20)
        self.description_entry = Entry(self.frame6_admin_page, font=20)

        # Place Entries
        self.amount_paid_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.category_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.description_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)

        # Create and Place register button
        self.register_expense = Button(self.frame7_admin_page, text='Registrar Egreso', command=self.log_expense, bg=self.ORANGE, fg=self.WHITE)
        self.register_expense.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)

    def log_expense(self):
        pass

    def remove_product_form(self):
        self.remove_everything()
        # Create Frames
        self.frame1_admin_page.place_forget()
        self.frame2_admin_page.place_forget()
        self.frame1_admin_page.place(relx=self.width2_1 - .15, rely=self.height2_1 - 0.05, relwidth=0.3, relheight=.1)
        self.frame2_admin_page.place(relx=self.width2_2 - .15, rely=self.height2_1 - 0.05, relwidth=0.3, relheight=.1)
        self.frame3_admin_page = Frame(root, bg=self.GREY)
        self.frame3_admin_page.place(relx=self.width2_1 - .15, rely=self.height2_2 - .05, relwidth=.3, relheight=.1)
        self.frame4_admin_page = Frame(root, bg=self.GREY)
        self.frame4_admin_page.place(relx=self.width2_2 - .15, rely=self.height2_2 - .05, relwidth=.3, relheight=.1)
        # Create Listbox and Button
        self.listbox_of_products_id = Listbox(self.frame3_admin_page)
        counter = 0
        for product_information in self.complete_products_data:
            counter += 1
            self.listbox_of_products_id.insert(counter, product_information[1])
        self.listbox_of_products_id.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        self.remove_product_button = Button(self.frame4_admin_page, text="Quitar Producto", command=self.remove_product, bg=self.ORANGE, fg=self.WHITE)
        self.remove_product_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def remove_product(self):
        product = self.listbox_of_products_id.get(self.listbox_of_products_id.curselection())
        try:
            for p in range(len(self.complete_products_data)):
                if product == self.complete_products_data[p][1]:
                    self.complete_products_data.remove(self.complete_products_data[p])
        except IndexError:
            pass
        json.dump(self.complete_products_data, open('products_information.txt', 'w'))
        self.remove_product_form()

    def inventory_list(self):
        # Remove 'Añadir Inverntario'
        self.remove_everything()
        # Frames
        self.frame1_admin_page.place_forget()
        self.frame2_admin_page.place_forget()
        self.frame1_admin_page.place(relx=self.width2_1 - .15, rely=self.height3_1 - .05, relwidth=.3, relheight=.1)
        self.frame2_admin_page.place(relx=self.width2_2 - .15, rely=self.height3_1 - .05, relwidth=.3, relheight=.1)
        self.frame3_admin_page = Frame(root, bg=self.GREY)
        self.frame3_admin_page.place(relx=self.width4_1 - .175, rely=self.height3_2 - .05, relwidth=.2, relheight=.1)
        self.frame4_admin_page = Frame(root, bg=self.GREY)
        self.frame4_admin_page.place(relx=self.width4_2 - .125, rely=self.height3_2 - .05, relwidth=.2, relheight=.1)
        self.frame5_admin_page = Frame(root, bg=self.GREY)
        self.frame5_admin_page.place(relx=self.width4_3 - .075, rely=self.height3_2 - .05, relwidth=.2, relheight=.1)
        self.frame6_admin_page = Frame(root, bg=self.GREY)
        self.frame6_admin_page.place(relx=self.width4_4 - .025, rely=self.height3_2 - .05, relwidth=.2, relheight=.1)
        self.frame7_admin_page = Frame(root, bg=self.GREY)
        self.frame7_admin_page.place(relx=self.width1_1-.15, rely=self.height3_3-.04, relwidth=.3, relheight=.08)
        # Listbox of products and button
        self.listbox_of_products_id = Listbox(self.frame3_admin_page)
        self.listbox_of_products = Listbox(self.frame4_admin_page)
        counter = 0
        for product_information in self.complete_products_data:
            counter += 1
            self.listbox_of_products.insert(counter, product_information[1])
            self.listbox_of_products_id.insert(counter, product_information[0])
        self.listbox_of_products.bind('<<ListboxSelect>>', lambda event, a=1: self.get_value1(a))
        self.listbox_of_products_id.bind('<<ListboxSelect>>', lambda event, a=0: self.get_value1(a))
        self.listbox_of_products.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)
        self.listbox_of_products_id.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        self.add_inventory_entry = Entry(self.frame5_admin_page, justify='center', font=20)
        self.add_inventory_entry.place(relx=0.05, rely=.05, relwidth=0.9, relheight=0.4)
        self.add_inventory_label = Label(self.frame5_admin_page, text='Cantidad', fg=self.WHITE)
        self.add_inventory_label.place(relx=.05, rely=.5, relwidth=0.9, relheight=0.4)
        self.payment_type = Listbox(self.frame6_admin_page, height=3)
        self.payment_type.place(relx=.05, rely=.1, relheight=.8, relwidth=.9)
        for i in range(len(self.payment_types)):
            self.payment_type.insert(i, self.payment_types[i])
        self.add_inventory_button = Button(self.frame7_admin_page, text="Registrar Inventario", command=self.register_inventory, bg=self.ORANGE, fg=self.WHITE)
        self.add_inventory_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def register_inventory(self):
        product = self.last_product
        for i in range(0, len(self.complete_products_data)):
            if product in self.complete_products_data[i]:
                self.complete_products_data[i][4] += int(self.add_inventory_entry.get())
                json.dump(self.complete_products_data, open('products_information.txt', 'w'))
                ws = self.wb['Egresos']
                # Get data ready
                current_date = datetime.now()
                current_date = f'{current_date.day}/{current_date.month}/{current_date.year}'
                payment_type = self.payment_type.get(self.payment_type.curselection())
                total_price = self.complete_products_data[i][5] * int(self.add_inventory_entry.get())
                product = self.complete_products_data[i][1]
                quantity = self.add_inventory_entry.get()
                ws.insert_rows(2)
                ws.cell(row=2, column=1, value=current_date)
                ws.cell(row=2, column=2, value=payment_type)
                ws.cell(row=2, column=3, value=total_price)
                ws.cell(row=2, column=4, value=product)
                ws.cell(row=2, column=5, value=quantity)
                self.wb.save('income_expenses.xlsx')
                self.wb = openpyxl.load_workbook('income_expenses.xlsx')

    def add_product_to_sell(self):
        self.remove_everything()
        # Frames
        self.frame1_admin_page.place_forget()
        self.frame2_admin_page.place_forget()
        self.frame1_admin_page.place(relx=self.width2_1 - 0.15, rely=self.height4_1 - 0.05, relwidth=0.3, relheight=.1)
        self.frame2_admin_page.place(relx=self.width2_2 - 0.15, rely=self.height4_1 - 0.05, relwidth=0.3, relheight=.1)
        self.frame3_admin_page = Frame(root, bg=self.GREY)
        self.frame3_admin_page.place(relx=self.width3_1 - 0.1, rely=self.height4_2 - 0.05, relwidth=0.2, relheight=.1)
        self.frame4_admin_page = Frame(root, bg=self.GREY)
        self.frame4_admin_page.place(relx=self.width3_2 - 0.1, rely=self.height4_2 - 0.05, relwidth=0.2, relheight=.1)
        self.frame5_admin_page = Frame(root, bg=self.GREY)
        self.frame5_admin_page.place(relx=self.width3_3 - 0.1, rely=self.height4_2 - 0.05, relwidth=0.2, relheight=.1)
        self.frame6_admin_page = Frame(root, bg=self.GREY)
        self.frame6_admin_page.place(relx=self.width3_1 - 0.1, rely=self.height4_3 - 0.05, relwidth=0.2, relheight=.1)
        self.frame7_admin_page = Frame(root, bg=self.GREY)
        self.frame7_admin_page.place(relx=self.width3_2 - 0.1, rely=self.height4_3 - 0.05, relwidth=0.2, relheight=.1)
        self.frame8_admin_page = Frame(root, bg=self.GREY)
        self.frame8_admin_page.place(relx=self.width3_3 - 0.1, rely=self.height4_3 - 0.05, relwidth=0.2, relheight=.1)
        self.frame9_admin_page = Frame(root, bg=self.GREY)
        self.frame9_admin_page.place(relx=self.width2_1 - 0.15, rely=self.height4_4 - 0.05, relwidth=0.3, relheight=.1)
        self.frame10_admin_page = Frame(root, bg=self.GREY)
        self.frame10_admin_page.place(relx=self.width2_2-.15, rely=self.height4_4-.05, relwidth=.3, relheight=.1)

        # Information Widgets
        self.item_id_label = Label(self.frame3_admin_page, bg=self.GREY, fg=self.WHITE, text="Codigo de Producto")
        self.item_id_label.place(relx=0.15, rely=0.5, relwidth=0.7, relheight=0.3)
        self.item_id_entry = Entry(self.frame3_admin_page, font=20)
        self.item_id_entry.place(relx=0.15, rely=0.08, relwidth=0.7, relheight=0.3)
        self.item_to_sell_label = Label(self.frame4_admin_page, bg=self.GREY, fg=self.WHITE, text="Nombre de Producto")
        self.item_to_sell_label.place(relx=0.15, rely=0.5, relwidth=0.7, relheight=0.3)
        self.item_to_sell_entry = Entry(self.frame4_admin_page, font=20)
        self.item_to_sell_entry.place(relx=0.15, rely=0.08, relwidth=0.7, relheight=0.3)
        self.item_provider_label = Label(self.frame5_admin_page, bg=self.GREY, fg=self.WHITE, text="Proveedor")
        self.item_provider_label.place(relx=0.15, rely=0.5, relwidth=0.7, relheight=0.3)
        self.item_provider_entry = Entry(self.frame5_admin_page, font=20)
        self.item_provider_entry.place(relx=0.15, rely=0.08, relwidth=0.7, relheight=0.3)
        self.price_label = Label(self.frame6_admin_page, bg=self.GREY, fg=self.WHITE, text="Precio de Producto")
        self.price_label.place(relx=0.15, rely=0.5, relwidth=0.7, relheight=0.3)
        self.price_entry = Entry(self.frame6_admin_page, font=20)
        self.price_entry.place(relx=0.15, rely=0.08, relwidth=0.7, relheight=0.3)
        self.current_inventory_label = Label(self.frame7_admin_page, bg=self.GREY, fg=self.WHITE, text="Cantidad en Inventario")
        self.current_inventory_label.place(relx=0.15, rely=0.5, relwidth=0.7, relheight=0.3)
        self.current_inventory_entry = Entry(self.frame7_admin_page, font=20)
        self.current_inventory_entry.place(relx=0.15, rely=0.08, relwidth=0.7, relheight=0.3)
        self.cost_entry = Entry(self.frame8_admin_page, font=20)
        self.cost_entry.place(relx=0.15, rely=0.08, relwidth=0.7, relheight=0.3)
        self.cost_label = Label(self.frame8_admin_page, bg=self.GREY, fg=self.WHITE, text='Costo de Producto')
        self.cost_label.place(relx=0.15, rely=0.5, relwidth=0.7, relheight=0.3)
        self.payment_type = Listbox(self.frame9_admin_page, height=3)
        self.payment_type.place(relx=.05, rely=.1, relheight=.8, relwidth=.9)
        for i in range(len(self.payment_types)):
            self.payment_type.insert(i, self.payment_types[i])
        # Register Widget
        self.register_button = Button(self.frame10_admin_page, text="Registrar Producto", bg=self.ORANGE, fg=self.WHITE, command=self.update_products_information)
        self.register_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def update_products_information(self):
        temp_data = [self.item_id_entry.get(), self.item_to_sell_entry.get(), self.item_provider_entry.get(), int(self.price_entry.get()), int(self.current_inventory_entry.get()), int(self.cost_entry.get())]
        if temp_data not in self.complete_products_data:
            self.complete_products_data.append(temp_data)
            json.dump(self.complete_products_data, open('products_information.txt', 'w'))
            ws = self.wb['Egresos']
            current_date = datetime.now()
            current_date = f'{current_date.day}/{current_date.month}/{current_date.year}'
            payment_type = self.payment_type.get(self.payment_type.curselection())
            total_price = int(self.cost_entry.get()) * int(self.current_inventory_entry.get())
            product = self.item_to_sell_entry.get()
            quantity = self.current_inventory_entry.get()
            ws.insert_rows(2)
            ws.cell(row=2, column=1, value=current_date)
            ws.cell(row=2, column=2, value=payment_type)
            ws.cell(row=2, column=3, value=total_price)
            ws.cell(row=2, column=4, value=product)
            ws.cell(row=2, column=5, value=quantity)
            self.wb.save('income_expenses.xlsx')
            self.wb = openpyxl.load_workbook('income_expenses.xlsx')

    def add_class_to_sell(self):
        self.remove_everything()
        # Frames
        self.frame1_admin_page.place_forget()
        self.frame2_admin_page.place_forget()
        self.frame1_admin_page.place(relx=self.width2_1-.15, rely=self.height4_1-.05, relheight=.1, relwidth=.3)
        self.frame2_admin_page.place(relx=self.width2_2-.15, rely=self.height4_1-.05, relheight=.1, relwidth=.3)
        self.frame3_admin_page = Frame(root, bg=self.GREY)
        self.frame3_admin_page.place(relx=self.width2_1-.15, rely=self.height4_2-.04, relheight=.08, relwidth=.3)
        self.frame4_admin_page = Frame(root, bg=self.GREY)
        self.frame4_admin_page.place(relx=self.width2_2-.15, rely=self.height4_2-.04, relheight=.08, relwidth=.3)
        # Añadir Actividad
        self.activity_button = Button(self.frame3_admin_page, text="Añadir Actividad", bg=self.ORANGE, fg=self.WHITE, command=self.add_activity)
        self.activity_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        # Añadir Horario, Dia, Maestro
        self.add_class_button = Button(self.frame4_admin_page, text='Añadir Clase', bg=self.ORANGE, fg=self.WHITE, command=self.check_activity)
        self.add_class_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def add_activity(self):
        self.remove_everything()
        # Frames
        self.frame1_activity_page = Frame(root, bg=self.GREY)
        self.frame1_activity_page.place(relx=self.width3_1-.1, rely=self.height4_2-.05, relwidth=.2, relheight=.1)
        self.frame2_activity_page = Frame(root, bg=self.GREY)
        self.frame2_activity_page.place(relx=self.width3_2-.1, rely=self.height4_2-.05, relwidth=.2, relheight=.1)
        self.frame3_activity_page = Frame(root, bg=self.GREY)
        self.frame3_activity_page.place(relx=self.width3_3-.1, rely= self.height4_2-.05, relwidth=.2, relheight=.1)
        # Add Activity Entry and Button
        self.add_activity_entry = Entry(self.frame1_activity_page, font=20)
        self.add_activity_entry.place(relx=0.05, rely=.075, relwidth=0.9, relheight=0.4)
        self.add_activity_label = Label(self.frame1_activity_page, bg=self.GREY, text='Actividad')
        self.add_activity_label.place(relx=0.05, rely=.5, relwidth=.9, relheight=.4)
        self.add_levels_entry = Entry(self.frame2_activity_page, fg='grey', font=20)
        self.add_levels_entry.place(relx=.05, rely=.075, relwidth=.9, relheight=.4)
        self.add_levels_entry.insert(0, 'Ejemplo: (1, 2, 3) o (blanco, amarillo, verde)')
        self.add_levels_entry.bind('<FocusIn>', self.focus_in1)
        self.add_levels_label = Label(self.frame2_activity_page, bg=self.GREY, text='Niveles')
        self.add_levels_label.place(relx=0.05, rely=.5, relwidth=.9, relheight=.4)
        self.add_activity_button = Button(self.frame3_activity_page, bg=self.ORANGE, fg=self.WHITE, text='Añadir Actividad', command=self.add_activity_to_sell)
        self.add_activity_button.place(relx=0.05, rely=.1, relwidth=.9, relheight=.8)

    def add_activity_to_sell(self):
        self.complete_classes_data.append([self.add_activity_entry.get(), self.add_levels_entry.get()])
        json.dump(self.complete_classes_data, open('Available_Classes.txt', 'w'))

    def check_activity(self):
        self.remove_everything()
        # Create Frames
        self.frame1_class_page = Frame(root, bg=self.GREY)
        self.frame1_class_page.place(relx=self.width3_1-.1, rely=self.height4_2-.05, relwidth=.2, relheight=.1)
        self.frame2_class_page = Frame(root, bg=self.GREY)
        self.frame2_class_page.place(relx=self.width3_2-.1, rely=self.height4_2-.05, relwidth=.2, relheight=.1)
        self.frame3_class_page = Frame(root, bg=self.GREY)
        self.frame3_class_page.place(relx=self.width3_3-.1, rely=self.height4_2-.05, relwidth=.2, relheight=.1)
        self.frame4_class_page = Frame(root, bg=self.GREY)
        self.frame4_class_page.place(relx=self.width2_1-.15, rely=self.height4_3-.05, relwidth=.3, relheight=.1)
        self.frame5_class_page = Frame(root, bg=self.GREY)
        self.frame5_class_page.place(relx=self.width2_2-.15, rely=self.height4_3-.05, relwidth=.3, relheight=.1)
        self.frame6_class_page = Frame(root, bg=self.GREY)
        self.frame6_class_page.place(relx=self.width2_1-.15, rely=self.height4_4 - .04, relwidth=.3, relheight=.08)
        self.frame7_class_page = Frame(root, bg=self.GREY)
        self.frame7_class_page.place(relx=self.width2_2-.15, rely=self.height4_4 - .04, relwidth=.3, relheight=.08)
        # Create Listbox, Entries, Labels, and Button
        self.list_of_activities = Combobox(self.frame1_class_page)
        self.list_of_activities.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)
        self.day_of_week = Listbox(self.frame2_class_page, height=4)
        self.day_of_week.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)
        self.list_of_activities.bind('<Return>', self.get_value)

        # Create Time Schedule example Entry
        self.time_schedule_entry = Entry(self.frame4_class_page, fg='grey', font=20)
        self.time_schedule_entry.place(relx=.05, rely=.05, relwidth=.9, relheight=.4)
        self.time_schedule_entry.insert(0, 'Ejemplo (12:00-14:00)')
        self.time_schedule_entry.bind('<FocusIn>', self.focus_in)

        # Create rest of widgets
        self.time_schedule_label = Label(self.frame4_class_page, bg=self.GREY, fg=self.WHITE, text='Horario')
        self.time_schedule_label.place(relx=.05, rely=.5, relwidth=.9, relheight=.4)
        self.teacher_entry = Entry(self.frame5_class_page, font=20)
        self.teacher_entry.place(relx=.05, rely=.05, relwidth=.9, relheight=.4)
        self.teacher_label = Label(self.frame5_class_page, bg=self.GREY, fg=self.WHITE, text='Maestro')
        self.teacher_label.place(relx=.05, rely=.5, relwidth=.9, relheight=.4)
        self.available_seats_entry = Entry(self.frame6_class_page, font=20)
        self.available_seats_entry.place(relx=.05, rely=.05, relwidth=.9, relheight=.4)
        self.available_seats_label = Label(self.frame6_class_page, bg=self.GREY, fg=self.WHITE, text='Lugares Disponibles')
        self.available_seats_label.place(relx=.05, rely=.5, relwidth=.9, relheight=.4)
        self.list_of_activities_button = Button(self.frame7_class_page, bg=self.ORANGE, fg=self.WHITE, text='Registrar Clase', command=self.add_class)
        self.list_of_activities_button.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)
        # Add days of week
        self.day_of_week.insert(0, 'Lunes')
        self.day_of_week.insert(1, 'Martes')
        self.day_of_week.insert(2, 'Miercoles')
        self.day_of_week.insert(3, 'Jueves')
        self.day_of_week.insert(4, 'Viernes')
        self.day_of_week.insert(5, 'Sabado')
        # List activities
        list_to_tupple = []
        for i in self.complete_classes_data:
            list_to_tupple.append(i[0])
        self.list_of_activities['values'] = tuple(list_to_tupple)

    def add_class(self):
        activity = self.list_of_activities.get()
        day_of_week = self.day_of_week.get(self.day_of_week.curselection())
        start_time, end_time = self.time_schedule_entry.get().split('-')
        start_hour, start_minute = start_time.split(':')
        end_hour, end_minute = end_time.split(':')
        start_hour, start_minute, end_hour, end_minute = int(start_hour), int(start_minute), int(end_hour), int(end_minute)
        levels = []
        for i in self.frame3_class_page.winfo_children():
            if i.instate(['selected']):
                levels.append(i['text'])
        if start_hour + 1 == end_hour:
            ready_to_append = [levels, day_of_week, self.time_schedule_entry.get(), self.teacher_entry.get(), int(self.available_seats_entry.get())]
            for i in range(len(self.complete_classes_data)):
                if self.complete_classes_data[i][0] == activity:
                    self.complete_classes_data[i].append(ready_to_append)
        else:
            while start_hour != end_hour:
                for i in range(len(self.complete_classes_data)):
                    if self.complete_classes_data[i][0] == activity:
                        self.complete_classes_data[i].append([levels, day_of_week, f'{start_hour}:{"00" if start_minute == 0 else start_minute}-{start_hour+1}:{"00" if end_minute == 0 else end_minute}', self.teacher_entry.get(), int(self.available_seats_entry.get())])
                        start_hour += 1
        json.dump(self.complete_classes_data, open('Available_Classes.txt', 'w'))

    def sell_mode(self):
        try:
            self.return_button.destroy()
        except AttributeError:
            pass
        try:
            self.admin_mode.destroy()
            self.seller_mode.destroy()
            self.frame1_sell.destroy()
            self.frame2_sell.destroy()
        except AttributeError:
            pass
        # Create Frames
        self.frame1_seller = Frame(root, bg=self.GREY)
        self.frame1_seller.place(relx=self.width2_1 - 0.15, rely=self.height1_1 - 0.04, relwidth=0.3, relheight=.08)
        self.frame2_seller = Frame(root, bg=self.GREY)
        self.frame2_seller.place(relx=self.width2_2 - 0.15, rely=self.height1_1 - 0.04, relwidth=0.3, relheight=.08)
        # Main Widgets
        self.return_button = Button(root, text="Regresar", command=self.main_page, bg=self.ORANGE, fg=self.WHITE)
        self.return_button.place(relwidth=0.13, relheight=0.06)
        self.class_button = Button(self.frame1_seller, text="Clases", command=self.classes, bg=self.ORANGE, fg=self.WHITE)
        self.class_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        self.product_button = Button(self.frame2_seller, text="Productos", command=self.products, bg=self.ORANGE, fg=self.WHITE)
        self.product_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)

    def products(self):
        self.remove_everything()
        # frames
        self.register_products_frame = Frame(root)
        self.divide_frame = Frame(root, bg=self.GREY)
        self.products_frame = Frame(root)
        self.register_products_frame.place(relx=0, rely=0.16, relwidth=0.2, relheight=0.84)
        self.divide_frame.place(relx=0.2, rely=0, relwidth=.04, relheight=1)
        self.products_frame.place(relx=0.24, rely=0, relwidth=0.76, relheight=1)
        self.buttons_frame = Frame(root)
        self.buttons_frame.place(relx=0, rely=0, relwidth=.2, relheight=0.16)
        # Create return button
        self.return_button = Button(self.buttons_frame, text="Regresar", command=self.clear_sell, bg=self.ORANGE, fg=self.WHITE)
        self.return_button.grid(row=0, column=0, padx=5, pady=5)
        self.charge_button = Button(self.buttons_frame, text="Cobrar", command=self.charge, bg=self.ORANGE, fg=self.WHITE)
        self.charge_button.grid(row=0, column=1, padx=5, pady=5)
        self.clear_button = Button(self.buttons_frame, text="Vaciar", command=self.clear, bg=self.ORANGE, fg=self.WHITE)
        self.clear_button.grid(row=0, column=2, padx=5, pady=5)
        self.payment_type = Listbox(self.buttons_frame, width=15, height=3)
        self.payment_type.bind('<<ListboxSelect>>', self.get_value2)
        self.payment_type.grid(row=1, column=0, padx=5, pady=5, columnspan=3)
        for i in range(len(self.payment_types)):
            self.payment_type.insert(i, self.payment_types[i])
        # Row and column
        counter = 0
        row = 0
        # Counter
        counter2 = 0
        for i in self.complete_products_data:
            Frame(self.products_frame, bg=self.GREY).place(relx=counter * .13 + counter * .005 + .005, rely=row * .08 + row * .005 + .005, relwidth=.13, relheight=.08)

            counter += 1
            if counter == 7:
                counter = 0
                row += 1
        for i in self.products_frame.winfo_children():
            # If category
                # Create Combo box
                # Create plus and minus buttons below
                # Make plus and minus buttons get combobox text and add label
            # If not category
            Label(i, text=self.complete_products_data[counter2][1], fg=self.WHITE).place(relx=0.1, rely=.05, relwidth=.8, relheight=.45)
            Button(i, text='+', font=15, bg=self.ORANGE, command=lambda m=self.complete_products_data[counter2][1]: self.add_product_label(m)).place(relx=0.25, rely=.55, relwidth=.2, relheight=.4)
            Button(i, text='-', font=15, bg=self.ORANGE, command=lambda m=self.complete_products_data[counter2][1]: self.remove_product_label(m)).place(relx=.55, rely=.55, relwidth=.2, relheight=.4)
            counter2 += 1

    def clear_sell(self):
        # Remove 'Productos' In sell mode
        for i in root.winfo_children():
            i.destroy()
        self.main_page()

    def remove_product_label(self, product):
        row = 1
        self.total_price = 0
        price = 0
        for i in self.register_products_frame.winfo_children():
            i.destroy()

        for i in self.list_of_labels:
            if i[1] == 0:
                pass
            if i[0] == product:
                i[1] -= 1
                if i[1] == 0:
                    self.list_of_labels.remove(i)

        for i in self.list_of_labels:
            for p in self.complete_products_data:
                if p[1] == i[0]:
                    price = p[3] * i[1]
            Label(self.register_products_frame, text=f'Producto: {i[0]}, Cantidad: {i[1]}, Precio: {price}', font=('Times', 14)).grid(row=row, column=0)
            self.total_price += price
            row += 1
        if self.total_price != 0:
            Label(self.register_products_frame, text=f'Total: {self.total_price}').grid(row=row + 1, column=0)

    def add_product_label(self, product):
        counter = 0
        row = 1
        self.total_price = 0
        price = 0
        for i in self.register_products_frame.winfo_children():
            if i['text'] == 'Regresar' or i['text'] == 'Cobrar':
                pass
            else:
                i.destroy()
        for i in self.list_of_labels:
            for p in self.complete_products_data:
                if p[1] == i[0] and p[4] > i[1]:
                    if i[0] == product:
                        i[1] += 1
                        counter += 1
                elif p[1] == i[0] and p[4] == i[1] and i[0] == product:
                    counter += 1
                    messagebox.showinfo('Inventario', 'Se acabo este producto.')
        if counter == 0:
            for p in self.complete_products_data:
                if p[1] == product and p[4] == 0:
                    messagebox.showinfo('Inventario', 'Se acabo este producto.')
                elif p[1] == product and p[4] != 0:
                    self.list_of_labels.append([product, 1])
        for i in self.list_of_labels:
            for p in self.complete_products_data:
                if p[1] == i[0]:
                    price = p[3] * i[1]
            Label(self.register_products_frame, text=f'Producto: {i[0]}, Cantidad: {i[1]}, Precio: {price}', font=('Times', 14)).grid(row=row, column=0)
            self.total_price += price
            row += 1
        Label(self.register_products_frame, text=f'Total: {self.total_price}').grid(row=row+1, column=0)

    def charge(self):
        payment_type = self.payment_type
        if payment_type != 0:
            for i in self.list_of_labels:
                for p in self.complete_products_data:
                    if i[0] == p[1]:
                        p[4] -= i[1]
                        json.dump(self.complete_products_data, open('products_information.txt', 'w'))
            for i in self.register_products_frame.winfo_children():
                if 'Total' not in i['text']:
                    i.destroy()
                else:
                    # Date, payment type, income, products, quantity
                    current_date = datetime.now()
                    current_date = f'{current_date.day}/{current_date.month}/{current_date.year}'
                    payment_type = self.selected_payment_type
                    total = i['text']
                    total = total.replace('Total: ', '')
                    products = []
                    quantities = []
                    products_string = ''
                    quantities_string = ''
                    for p in self.list_of_labels:
                        products.append(p[0])
                        quantities.append(p[1])

                    products_string += products[0]
                    quantities_string += quantities[0]
                    for p in range(1, len(products)):
                        products_string += ', ' + products[p]
                        quantities_string += ', ' + quantities[p]
                    ws = self.wb['Ingresos']
                    ws.insert_rows(2)
                    ws.cell(row=2, column=1, value=current_date)
                    ws.cell(row=2, column=2, value=payment_type)
                    ws.cell(row=2, column=3, value=total)
                    ws.cell(row=2, column=4, value=products_string)
                    ws.cell(row=2, column=5, value=quantities_string)
                    self.wb.save('income_expenses.xlsx')
                    self.wb = openpyxl.load_workbook('income_expenses.xlsx')
                self.list_of_labels.clear()
                self.payment_type = 0

    def clear(self):
        for i in self.register_products_frame.winfo_children():
            i.destroy()
            self.list_of_labels.clear()
            self.payment_type = 0

    def classes(self):
        self.remove_everything()
        # Create Frames
        self.class_frame1 = Frame(root, bg=self.GREY)
        self.class_frame1.place(relx=self.width2_1-.15, rely=self.height1_1-.04, relwidth=.3, relheight=.08)
        self.class_frame2 = Frame(root, bg=self.GREY)
        self.class_frame2.place(relx=self.width2_2-.15, rely=self.height1_1-.04, relwidth=.3, relheight=.08)
        # Create Add and Remove Student buttons
        self.add_student_button = Button(self.class_frame1, text="Añadir Estudiante", command=self.add_student, bg=self.ORANGE, fg=self.WHITE)
        self.add_student_button.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)
        self.remove_student_button = Button(self.class_frame2, text='Quitar Estudiante', command=self.remove_student, bg=self.ORANGE, fg=self.WHITE)
        self.remove_student_button.place(relx=.05, rely=.1, relwidth=.9, relheight=.8)

    def add_student(self):
        self.remove_everything()
        self.return_button.destroy()
        # Create 16 frames
        self.frame1_add_student = Frame(root)
        self.frame2_add_student = Frame(root)
        self.frame3_add_student = Frame(root)
        self.frame4_add_student = Frame(root)
        self.frame5_add_student = Frame(root)
        self.frame6_add_student = Frame(root)
        self.frame7_add_student = Frame(root)
        self.frame8_add_student = Frame(root)
        self.frame9_add_student = Frame(root)
        self.frame10_add_student = Frame(root)
        self.frame11_add_student = Frame(root)
        self.frame12_add_student = Frame(root)
        self.frame13_add_student = Frame(root)
        self.frame14_add_student = Frame(root)
        self.frame15_add_student = Frame(root)
        self.frame16_add_student = Frame(root)
        # Place 16 frames
        counter = 0
        row = 0
        for i in root.winfo_children():
            i.place(relx=counter * .165 + .005, rely=row * .105 + .005, relwidth=.16, relheight=.08)
            counter += 1
            if counter == 6:
                counter = 0
                row += 1
        # Place return button
        self.return_button = Button(self.frame1_add_student, font=('Times', 20), text="Regresar", command=self.remove_user_information_interface, bg=self.ORANGE, fg=self.WHITE)
        self.return_button.place(relx=0.05, rely=.1, relwidth=0.9, relheight=0.8)
        # Place User Questions
        self.user_id_label = Label(self.frame2_add_student, font=('Times', 15), text='ID (Buscar por Numero)')
        self.user_id_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.user_id_entry = Entry(self.frame2_add_student, font=20)
        self.user_id_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.user_name_label = Label(self.frame3_add_student, font=('Times', 15), text='Nombre Completo Alumno*')
        self.user_name_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.user_name_entry = Entry(self.frame3_add_student, font=20)
        self.user_name_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.user_birthday_label = Label(self.frame4_add_student, font=('Times', 15), text='Fecha de Nacimiento*')
        self.user_birthday_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.user_birthday_entry = Entry(self.frame4_add_student, font=20)
        self.user_birthday_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.parent_address_label = Label(self.frame5_add_student, font=('Times', 15), text='Direccion')
        self.parent_address_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.parent_address_entry = Entry(self.frame5_add_student, font=20)
        self.parent_address_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.parent_phone_label = Label(self.frame6_add_student, font=('Times', 15), text='Telefono*')
        self.parent_phone_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.parent_phone_entry = Entry(self.frame6_add_student, font=20)
        self.parent_phone_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.parent_email_label = Label(self.frame7_add_student, font=('Times', 15), text='E-mail')
        self.parent_email_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.parent_email_entry = Entry(self.frame7_add_student, font=20)
        self.parent_email_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.father_name_label = Label(self.frame8_add_student, font=('Times', 15), text='Nombre Padre')
        self.father_name_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.father_name_entry = Entry(self.frame8_add_student, font=20)
        self.father_name_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.mother_name_label = Label(self.frame9_add_student, font=('Times', 15), text='Nombre Madre')
        self.mother_name_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.mother_name_entry = Entry(self.frame9_add_student, font=20)
        self.mother_name_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.inscription_date_label = Label(self.frame10_add_student, font=('Times', 15), text='Fecha de Inscripcion')
        self.inscription_date_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.inscription_date_entry = Entry(self.frame10_add_student, font=20)
        self.inscription_date_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.payment_day_label = Label(self.frame11_add_student, font=('Times', 15), text='Dia de Pago')
        self.payment_day_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.payment_day_entry = Entry(self.frame11_add_student, font=20)
        self.payment_day_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.amount_paid_label = Label(self.frame12_add_student, font=('Times', 15), text='Cantidad Pagada')
        self.amount_paid_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.amount_paid_entry = Entry(self.frame12_add_student, font=20)
        self.amount_paid_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.hours_of_access_label = Label(self.frame13_add_student, font=('Times', 15), text='Accesso a Horas')
        self.hours_of_access_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.hours_of_access_entry = Entry(self.frame13_add_student, font=20)
        self.hours_of_access_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.security_link_label = Label(self.frame14_add_student, font=('Times', 15), text='Link de Seguro')
        self.security_link_label.place(relheight=.4, relwidth=.9, relx=.05, rely=.5)
        self.security_link_entry = Entry(self.frame14_add_student, font=20)
        self.security_link_entry.place(relheight=.4, relwidth=.9, relx=.05, rely=.05)
        self.payment_type = Listbox(self.frame15_add_student, height=3)
        self.payment_type.bind('<<ListboxSelect>>', self.get_value2)
        self.payment_type.place(relheight=.9, relwidth=.9, relx=.05, rely=.05)
        for i in range(len(self.payment_types)):
            self.payment_type.insert(i, self.payment_types[i])
        self.register_user_information = Button(self.frame16_add_student, text='Registrar/Actualizar Informacion', command=self.register_user, font=('Times', 15))
        self.register_user_information.place(relheight=.8, relwidth=.9, rely=.1, relx=.05)

    def remove_user_information_interface(self):
        # Iterate through all frames
        for i in root.winfo_children():
            i.destroy()  # Destroy current Frame iteration
        self.main_page()

    def register_user(self):
        # Check if required entries are registered
        if self.user_birthday_entry.get() == '' or self.user_name_entry.get() == '' or self.parent_phone_entry.get() == '' or self.payment_type == 0:
            # Check ID Entry and Full Name Entry for existing users
            if self.check_records(self.user_id_entry.get(), self.user_name_entry.get()) != 1:
                # Create Label to present user with 'Fill in Required fields'
                Label(root, text='Llena los campos con *').place(relheight=.1, relwidth=.2, relx=.4, rely=.45)

        # Check if all required fields are filled
        else:
            user_id = self.user_id_entry.get()
            user_name = self.user_name_entry.get()
            # If there is no assigned user id create one
            if user_id == '':
                user_id = len(self.complete_students_data)
                while True:
                    counter = 0
                    # Format it as a string
                    if 0 <= user_id < 10:
                        user_id = f'000{user_id}'
                    elif 10 <= user_id < 100:
                        user_id = f'00{user_id}'
                    elif 100 <= user_id < 1000:
                        user_id = f'0{user_id}'
                    else:
                        user_id = str(user_id)
                    # Check if it is occupied
                    for i in self.complete_students_data:
                        if i[0] == user_id:
                            user_id = int(user_id) + 1
                        else:
                            counter += 1
                    if counter == len(self.complete_students_data):
                        break
            # Get the rest of the information
            birthday = self.user_birthday_entry.get()
            parent_address = self.parent_address_entry.get()
            parent_telephone = self.parent_phone_entry.get()
            parent_email = self.parent_email_entry.get()
            father_name = self.father_name_entry.get()
            mother_name = self.mother_name_entry.get()
            inscription_date = self.inscription_date_entry.get()
            payment_date = self.payment_day_entry.get()
            amount_paid = self.amount_paid_entry.get()
            amount_of_hours = self.hours_of_access_entry.get()
            security_link = self.security_link_entry.get()
            counter = 0
            counter2 = 0
            while counter < len(self.complete_students_data):
                if user_id == self.complete_students_data[counter][0]:
                    # New Payment Date
                    if self.complete_students_data[counter][8] != inscription_date:
                        current_date = inscription_date
                        payment_type = self.selected_payment_type
                        income = amount_paid
                        payment_type2 = 'Inscripcion'
                        description = f'Inscripcion de {user_name}'
                        ws = self.wb['Ingresos']
                        ws.insert_rows(2)
                        ws.cell(row=2, column=1, value=current_date)
                        ws.cell(row=2, column=2, value=payment_type)
                        ws.cell(row=2, column=3, value=income)
                        ws.cell(row=2, column=4, value=payment_type2)
                        ws.cell(row=2, column=5, value=description)
                        self.wb.save('income_expenses.xlsx')
                        self.wb = openpyxl.load_workbook('income_expenses.xlsx')
                    elif self.complete_students_data[counter][9] != payment_date:
                        current_date = payment_date
                        payment_type = self.selected_payment_type
                        income = amount_paid
                        payment_type2 = 'Mensualidad'
                        description = f'Mensualidad de {user_name}'
                        ws = self.wb['Ingresos']
                        ws.insert_rows(2)
                        ws.cell(row=2, column=1, value=current_date)
                        ws.cell(row=2, column=2, value=payment_type)
                        ws.cell(row=2, column=3, value=income)
                        ws.cell(row=2, column=4, value=payment_type2)
                        ws.cell(row=2, column=5, value=description)
                        self.wb.save('income_expenses.xlsx')
                        self.wb = openpyxl.load_workbook('income_expenses.xlsx')
                    self.complete_students_data.pop(counter)
                    counter2 = 1
                counter += 1
            # Check if it is the first payment
            if counter2 == 0 and payment_date != '':
                current_date = payment_date
                payment_type = self.selected_payment_type
                income = amount_paid
                payment_type2 = 'Mensualidad'
                description = f'Mensualidad de {user_name}'
                ws = self.wb['Ingresos']
                ws.insert_rows(2)
                ws.cell(row=2, column=1, value=current_date)
                ws.cell(row=2, column=2, value=payment_type)
                ws.cell(row=2, column=3, value=income)
                ws.cell(row=2, column=4, value=payment_type2)
                ws.cell(row=2, column=5, value=description)
                self.wb.save('income_expenses.xlsx')
                self.wb = openpyxl.load_workbook('income_expenses.xlsx')
            elif counter == 0 and inscription_date != '':
                current_date = inscription_date
                payment_type = self.selected_payment_type
                income = amount_paid
                payment_type2 = 'Inscripcion'
                description = f'Inscripcion de {user_name}'
                ws = self.wb['Ingresos']
                ws.insert_rows(2)
                ws.cell(row=2, column=1, value=current_date)
                ws.cell(row=2, column=2, value=payment_type)
                ws.cell(row=2, column=3, value=income)
                ws.cell(row=2, column=4, value=payment_type2)
                ws.cell(row=2, column=5, value=description)
                self.wb.save('income_expenses.xlsx')
                self.wb = openpyxl.load_workbook('income_expenses.xlsx')
            self.complete_students_data.append([user_id, user_name, birthday, parent_address, parent_telephone, parent_email, father_name, mother_name, inscription_date, payment_date, amount_paid, amount_of_hours, security_link])
            json.dump(self.complete_students_data, open('Students.txt', 'w'))
            # Create Frame for main grid
            self.frame16_add_student = Frame(root, bg='#9fb5b7')
            self.frame16_add_student.place(rely=.35, relx=.025, relheight=.6, relwidth=.85)
            # Create Grid
            Label(self.frame16_add_student, text='Dia', bg='#9fb5b7', font=('Times', 20)).grid(row=1, column=0, padx=20, pady=35)
            Label(self.frame16_add_student, text='Nivel', bg='#9fb5b7', font=('Times', 20)).grid(row=2, column=0, padx=20, pady=35)
            Label(self.frame16_add_student, text='Horario', bg='#9fb5b7', font=('Times', 20)).grid(row=3, column=0, padx=20, pady=35)
            Label(self.frame16_add_student, text='Profesor', bg='#9fb5b7', font=('Times', 20)).grid(row=4, column=0, padx=20, pady=35)
            Label(self.frame16_add_student, text='Cupo', bg='#9fb5b7', font=('Times', 20)).grid(row=5, column=0, padx=20, pady=35)
            column = 1
            day_of_week = ('L', 'M', 'M', 'J', 'V')
            for i in self.complete_classes_data:
                # Create list of activities
                Label(self.frame16_add_student, text=i[0], bg='#9fb5b7', font=('Times', 20)).grid(row=0, column=column, padx=20, pady=20)
                temp_frame = Frame(self.frame16_add_student, bg='#9fb5b7')

                # Create Checkboxes
                temp_frame.grid(row=1, column=column, padx=10, pady=35)
                for p in range(0, 5):
                    Checkbutton(temp_frame, text=day_of_week[p]).grid(row=0, column=p, sticky='s', padx=5)
                counter = 0

                # Create Listbox for levels
                temp_listbox = Listbox(self.frame16_add_student, height=3)
                temp_listbox.bind('<<ListboxSelect>>', lambda event, a=column: self.listbox_value(a))
                for p in i[1].split(', '):
                    temp_listbox.insert(counter, p)
                    counter += 1
                temp_listbox.grid(row=2, column=column, padx=10, pady=10)

                # Make Time Suggestions based on level and activity
                temp_combo_box = Combobox(self.frame16_add_student, font=('Times', 20))
                temp_combo_box.grid(row=3, column=column, padx=20, pady=35)

                # Professor Listbox
                temp_entry = Entry(self.frame16_add_student, font=('Times', 20))
                temp_entry.bind('<Return>', lambda event, a=column: self.get_seats(a))
                temp_entry.grid(row=4, column=column)

                # Create seats available
                Label(self.frame16_add_student, text='0', bg='#9fb5b7').grid(row=5, column=column)

                # Next Column
                column += 1

            # Create
            Button(root, text='Register Graph', command=self.register_graph).place(relx=.895, relheight=.1, relwidth=.095, rely=.6)

    def check_records(self, user_id, user_name):
        # Check if there is an ID and name
        if user_id != '' and user_name != '':
            for i in self.complete_students_data:
                if user_id == i[0]:
                    # Delete all Entries' information
                    self.user_birthday_entry.delete(0, END)
                    self.parent_address_entry.delete(0, END)
                    self.parent_phone_entry.delete(0, END)
                    self.parent_email_entry.delete(0, END)
                    self.father_name_entry.delete(0, END)
                    self.mother_name_entry.delete(0, END)
                    self.inscription_date_entry.delete(0, END)
                    self.payment_day_entry.delete(0, END)
                    self.amount_paid_entry.delete(0, END)
                    self.hours_of_access_entry.delete(0, END)
                    self.security_link_entry.delete(0, END)
                    # Get student information written into Entries
                    self.user_birthday_entry.insert(END, i[2])
                    self.parent_address_entry.insert(END, i[3])
                    self.parent_phone_entry.insert(END, i[4])
                    self.parent_email_entry.insert(END, i[5])
                    self.father_name_entry.insert(END, i[6])
                    self.mother_name_entry.insert(END, i[7])
                    self.inscription_date_entry.insert(END, i[8])
                    self.payment_day_entry.insert(END, i[9])
                    self.amount_paid_entry.insert(END, i[10])
                    self.hours_of_access_entry.insert(END, i[11])
                    self.security_link_entry.insert(END, i[12])
                    # Skip rest of register_user if it finds user id and name
                    return 1

        # Check for ID
        elif user_id != '':
            for i in self.complete_students_data:
                if user_id == i[0]:
                    # Delete all Entries' information
                    self.user_name_entry.delete(0, END)
                    self.user_birthday_entry.delete(0, END)
                    self.parent_address_entry.delete(0, END)
                    self.parent_phone_entry.delete(0, END)
                    self.parent_email_entry.delete(0, END)
                    self.father_name_entry.delete(0, END)
                    self.mother_name_entry.delete(0, END)
                    self.inscription_date_entry.delete(0, END)
                    self.payment_day_entry.delete(0, END)
                    self.amount_paid_entry.delete(0, END)
                    self.hours_of_access_entry.delete(0, END)
                    self.security_link_entry.delete(0, END)
                    # Get student information written into Entries
                    self.user_name_entry.insert(END, i[1])
                    self.user_birthday_entry.insert(END, i[2])
                    self.parent_address_entry.insert(END, i[3])
                    self.parent_phone_entry.insert(END, i[4])
                    self.parent_email_entry.insert(END, i[5])
                    self.father_name_entry.insert(END, i[6])
                    self.mother_name_entry.insert(END, i[7])
                    self.inscription_date_entry.insert(END, i[8])
                    self.payment_day_entry.insert(END, i[9])
                    self.amount_paid_entry.insert(END, i[10])
                    self.hours_of_access_entry.insert(END, i[11])
                    self.security_link_entry.insert(END, i[12])
                    # Skip rest of register_user if it finds user id
                    return 1
        # Check by Name
        elif user_name != '':
            for i in self.complete_students_data:
                if user_name == i[1]:
                    # Delete all Entries' information
                    self.user_id_entry.delete(0, END)
                    self.user_birthday_entry.delete(0, END)
                    self.parent_address_entry.delete(0, END)
                    self.parent_phone_entry.delete(0, END)
                    self.parent_email_entry.delete(0, END)
                    self.father_name_entry.delete(0, END)
                    self.mother_name_entry.delete(0, END)
                    self.inscription_date_entry.delete(0, END)
                    self.payment_day_entry.delete(0, END)
                    self.amount_paid_entry.delete(0, END)
                    self.hours_of_access_entry.delete(0, END)
                    self.security_link_entry.delete(0, END)
                    # Get student information written into Entries
                    self.user_id_entry.insert(END, i[0])
                    self.user_birthday_entry.insert(END, i[2])
                    self.parent_address_entry.insert(END, i[3])
                    self.parent_phone_entry.insert(END, i[4])
                    self.parent_email_entry.insert(END, i[5])
                    self.father_name_entry.insert(END, i[6])
                    self.mother_name_entry.insert(END, i[7])
                    self.inscription_date_entry.insert(END, i[8])
                    self.payment_day_entry.insert(END, i[9])
                    self.amount_paid_entry.insert(END, i[10])
                    self.hours_of_access_entry.insert(END, i[11])
                    self.security_link_entry.insert(END, i[12])
                    # Skip rest of register_user if it finds username
                    return 1

    def remove_student(self):
        self.remove_everything()

    # Gray out Text
    def focus_in(self, event):
        self.time_schedule_entry.delete(0, END)
        self.time_schedule_entry.config(fg='black')

    def focus_in1(self, event):
        self.add_levels_entry.delete(0, END)
        self.add_levels_entry.config(fg='black')

    def get_value(self, event):
        self.listbox_selection = self.list_of_activities.get()
        r = 0
        c = 0
        for i in self.complete_classes_data:
            if self.listbox_selection == i[0]:
                for p in i[1].split(', '):
                    Checkbutton(self.frame3_class_page, text=p).grid(row=r, column=c, padx=5, pady=5)
                    c += 1
                    if c == 3:
                        c = 0
                        r += 1

    def get_value1(self, column):
        try:
            if column == 0:
                self.last_product = self.listbox_of_products_id.get(self.listbox_of_products_id.curselection())
            if column == 1:
                self.last_product = self.listbox_of_products.get(self.listbox_of_products.curselection())
        except:
            pass

    def get_value2(self, event):
        try:
            self.selected_payment_type = self.payment_type.get(self.payment_type.curselection())
        except:
            pass

    def listbox_value(self, column):
        try:
            list_to_tuple = []
            for i in self.frame16_add_student.winfo_children():
                if i in self.frame16_add_student.grid_slaves(row=2, column=column):
                    self.last_selection = i.get(i.curselection())
            for x in self.frame16_add_student.winfo_children():
                if x in self.frame16_add_student.grid_slaves(row=0, column=column):
                    for i in self.complete_classes_data:
                        if x.cget('text') == i[0]:
                            for p in range(2, len(i)):
                                if self.last_selection in i[p][0]:
                                    list_to_tuple.append(i[p][2])

            for i in self.frame16_add_student.winfo_children():
                if i in self.frame16_add_student.grid_slaves(row=3, column=column):
                    i['values'] = tuple(list_to_tuple)
        except:
            pass

    def get_seats(self, column):
        days_of_week = ('Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes')
        for i in self.frame16_add_student.winfo_children():
            if i in self.frame16_add_student.grid_slaves(row=0, column=column):
                activity = i.cget('text')
            if i in self.frame16_add_student.grid_slaves(row=1, column=column):
                day_list = [p.instate(['selected']) for p in i.winfo_children()]
            if i in self.frame16_add_student.grid_slaves(row=2, column=column):
                level = self.last_selection
            if i in self.frame16_add_student.grid_slaves(row=3, column=column):
                times = []
                time = i.get()
                start_time, end_time = time.split('-')
                start_hour, start_minute = start_time.split(':')
                end_hour, end_minute = end_time.split(':')
                start_hour, end_hour = int(start_hour), int(end_hour)
                if start_hour + 1 == end_hour:
                    times.append(f'{start_hour}:{start_minute}-{start_hour+1}:{start_minute}')
                else:
                    while start_hour != end_hour:
                        times.append(f'{start_hour}:{start_minute}-{start_hour+1}:{start_minute}')
                        start_hour += 1

            if i in self.frame16_add_student.grid_slaves(row=4, column=column):
                teacher = i.get()

        for i in self.complete_classes_data:
            if activity == i[0]:
                for x in range(len(day_list)):
                    for y in times:
                        if day_list[x]:
                            for p in range(2, len(i)):
                                if level in i[p][0] and teacher in i[p] and days_of_week[x] in i[p] and y in i[p][2]:
                                    Label(self.frame16_add_student, text=i[p][4], bg='#9fb5b7').grid(row=5, column=column, padx=30)
    def register_graph(self):
        days_of_week = ('Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes')
        for column in range(1, len(self.complete_classes_data) + 1):
            not_empty = True
            for i in self.frame16_add_student.winfo_children():
                if i in self.frame16_add_student.grid_slaves(row=3, column=column):
                    try:
                        time = i.get()
                        start_time, end_time = time.split('-')
                    except ValueError:
                        not_empty = False
            if not_empty:
                for i in self.frame16_add_student.winfo_children():
                    if i in self.frame16_add_student.grid_slaves(row=0, column=column):
                        activity = i.cget('text')
                    if i in self.frame16_add_student.grid_slaves(row=1, column=column):
                        day_list = [p.instate(['selected']) for p in i.winfo_children()]
                    if i in self.frame16_add_student.grid_slaves(row=2, column=column):
                        level = self.last_selection
                    if i in self.frame16_add_student.grid_slaves(row=3, column=column):
                        times = []
                        time = i.get()
                        start_time, end_time = time.split('-')
                        start_hour, start_minute = start_time.split(':')
                        end_hour, end_minute = end_time.split(':')
                        start_hour, end_hour = int(start_hour), int(end_hour)
                        if start_hour + 1 == end_hour:
                            times.append(f'{start_hour}:{start_minute}-{start_hour + 1}:{start_minute}')
                        else:
                            while start_hour != end_hour:
                                times.append(f'{start_hour}:{start_minute}-{start_hour + 1}:{start_minute}')
                                start_hour += 1
                    if i in self.frame16_add_student.grid_slaves(row=4, column=column):
                        teacher = i.get()
                counter = 0
                for i in self.complete_classes_data:
                    if activity == i[0]:
                        for x in range(len(day_list)):
                            for y in times:
                                counter2 = 0
                                if day_list[x]:
                                    for p in range(2, len(i)):
                                        counter2 = 1 if (z for z in times) not in i[p] else 0
                                        if counter2 == 0 or days_of_week[x] not in i[p]:
                                            try:
                                                self.not_available.destroy()
                                            except AttributeError:
                                                pass
                                            self.not_available = Label(root, text='Dia o Hora no disponible', font=('Times', 13))
                                            self.not_available.place(relx=.895, rely=.5, relwidth=.095, relheight=.08)
                                            counter += 1
                                        else:
                                            try:
                                                self.not_available.destroy()
                                            except AttributeError:
                                                pass

                for i in self.complete_classes_data:
                    if activity == i[0]:
                        for x in range(len(day_list)):
                            for y in times:
                                if day_list[x]:
                                    for p in range(2, len(i)):
                                        if teacher in i[p] and days_of_week[x] in i[p] and y in i[p][2] and counter == 0:
                                            i[p][4] -= 1
                                            json.dump(self.complete_classes_data, open('Available_Classes.txt', 'w'))
                                            for z in range(len(self.complete_students_data)):
                                                if self.complete_students_data[z][0] == self.user_id_entry.get():
                                                    self.complete_students_data[z].append([activity, days_of_week[x], y, teacher])
                                                    json.dump(self.complete_students_data, open('Students.txt', 'w'))



if __name__ == '__main__':
    front = FrontEnd()
    front.sell_mode()
    root.mainloop()